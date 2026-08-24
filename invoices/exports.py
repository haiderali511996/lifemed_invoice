"""Taking a list of invoices out of the system, in one piece.

Chasing money means working off a list somebody can sort, print and hand to
whoever is making the calls. Everything here builds that list from the same
rows the screen shows, so an export and the page it came from can never
disagree.

One row per invoice line, with the invoice's own figures repeated alongside.
That shape suits a spreadsheet: it filters, it sorts, and it pivots. A row
per invoice would have had to bury the items in a cell.
"""

import csv
from decimal import Decimal
from io import BytesIO, StringIO

try:
    import pymupdf as fitz
except ImportError:  # pragma: no cover
    import fitz

ZERO = Decimal("0.00")

COLUMNS = [
    ("invoice_no", "Invoice #", 14),
    ("date", "Date", 11),
    ("customer", "Customer", 28),
    ("address", "Address", 34),
    ("license_no", "Licence #", 14),
    ("sales_rep", "Sales Rep", 18),
    ("item", "Item", 30),
    ("batch", "Batch", 12),
    ("expiry", "Expiry", 9),
    ("qty", "Qty", 7),
    ("bonus", "Bonus", 7),
    ("price", "T.Price", 11),
    ("discount", "Disc %", 8),
    ("line_total", "Line Amount", 13),
    ("invoice_total", "Invoice Total", 14),
    ("paid", "Paid", 12),
    ("credited", "Credited", 12),
    ("balance", "Balance", 13),
    ("status", "Status", 10),
    ("days", "Days", 7),
]

# Repeating the invoice's own totals against every one of its lines would add
# them up several times over, so only the first line of each invoice carries
# them and the rest are left blank - which is also how a human reads it.
INVOICE_LEVEL = {"invoice_total", "paid", "credited", "balance", "status", "days"}

MONEY_COLUMNS = {
    "price", "line_total", "invoice_total", "paid", "credited", "balance",
}


def _line_total(item):
    net = item.price - (item.price * item.discount / Decimal("100"))

    return (net * item.qty).quantize(ZERO)


def _status(invoice):
    if invoice.balance <= ZERO:
        return "Paid"

    return "Overdue" if invoice.is_overdue else "Unpaid"


def invoice_rows(invoices):
    """One row per invoice line, invoice figures on the first line of each.

    An invoice with no lines still appears: it was raised, it may be owed,
    and dropping it would quietly shorten what is being chased.
    """
    rows = []

    for invoice in invoices:
        customer = invoice.customer

        shared = {
            "invoice_no": invoice.invoice_no,
            "date": invoice.date,
            "customer": customer.name,
            "address": customer.address or "",
            "license_no": invoice.license_no or "",
            "sales_rep": (
                invoice.sales_rep.full_name if invoice.sales_rep else ""
            ),
            "invoice_total": invoice.total,
            "paid": invoice.amount_paid,
            "credited": invoice.amount_returned,
            "balance": invoice.balance,
            "status": _status(invoice),
            "days": invoice.days_outstanding,
        }

        items = list(invoice.items.all())

        if not items:
            rows.append({
                **shared,
                "first": True,
                "item": "(no items recorded)",
                "batch": "", "expiry": "",
                "qty": "", "bonus": "", "price": "", "discount": "",
                "line_total": "",
            })
            continue

        for position, item in enumerate(items):
            row = {
                **shared,
                # The spreadsheet keeps the customer on every row so it can be
                # filtered and pivoted; the document uses this to print them
                # once, which is how a person reads a list.
                "first": position == 0,
                "item": item.name,
                "batch": item.batch or "",
                "expiry": item.expiry or "",
                "qty": item.qty,
                "bonus": item.bonus or "",
                "price": item.price,
                "discount": item.discount,
                "line_total": _line_total(item),
            }

            if position:
                for field in INVOICE_LEVEL:
                    row[field] = ""

            rows.append(row)

    return rows


def totals_of(invoices):
    """The figures that belong under the list, counted once per invoice."""
    return {
        "invoices": len(invoices),
        "total": sum((i.total for i in invoices), ZERO),
        "paid": sum((i.amount_paid for i in invoices), ZERO),
        "credited": sum((i.amount_returned for i in invoices), ZERO),
        "balance": sum((i.balance for i in invoices), ZERO),
    }


# ------------------------------------------------------------------ spreadsheet

def to_csv(invoices):
    """A comma-separated file, written with a byte order mark.

    Excel reads a UTF-8 file without one as the system's legacy encoding and
    mangles every address that is not plain ASCII, so the mark is what makes
    this open cleanly by double-clicking it.
    """
    buffer = StringIO()
    writer = csv.writer(buffer)

    writer.writerow([label for _, label, _ in COLUMNS])

    for row in invoice_rows(invoices):
        writer.writerow([row[field] for field, _, _ in COLUMNS])

    figures = totals_of(invoices)

    writer.writerow([])
    writer.writerow([
        "TOTAL", f"{figures['invoices']} invoice(s)", "", "", "", "", "", "",
        "", "", "", "", "",
        "", figures["total"], figures["paid"], figures["credited"],
        figures["balance"], "", "",
    ])

    return buffer.getvalue().encode("utf-8-sig")


def to_xlsx(invoices):
    """A real spreadsheet, when openpyxl is installed.

    Returns None when it is not, so the caller can fall back rather than the
    page breaking: a missing optional library should cost a nicer file, not
    the export.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    book = Workbook()
    sheet = book.active
    sheet.title = "Invoices"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="14304D")

    sheet.append([label for _, label, _ in COLUMNS])

    for position, (_, _, width) in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=1, column=position)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

        sheet.column_dimensions[get_column_letter(position)].width = width

    for row in invoice_rows(invoices):
        sheet.append([row[field] for field, _, _ in COLUMNS])

    money_format = "#,##0.00"

    for position, (field, _, _) in enumerate(COLUMNS, start=1):
        if field not in MONEY_COLUMNS:
            continue

        for cell in sheet.iter_cols(
            min_col=position, max_col=position, min_row=2
        ):
            for one in cell:
                one.number_format = money_format

    figures = totals_of(invoices)

    sheet.append([])

    totals_row = [""] * len(COLUMNS)
    totals_row[0] = "TOTAL"
    totals_row[1] = f"{figures['invoices']} invoice(s)"

    for field, value in (
        ("invoice_total", figures["total"]),
        ("paid", figures["paid"]),
        ("credited", figures["credited"]),
        ("balance", figures["balance"]),
    ):
        totals_row[[c[0] for c in COLUMNS].index(field)] = value

    sheet.append(totals_row)

    last = sheet.max_row

    for position in range(1, len(COLUMNS) + 1):
        sheet.cell(row=last, column=position).font = Font(bold=True)

    # Frozen header and filters, because this list exists to be sorted and
    # worked through rather than read top to bottom.
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{last - 2}"

    stream = BytesIO()
    book.save(stream)

    return stream.getvalue()


# ------------------------------------------------------------------- document

PAGE = fitz.paper_rect("a4-l")

MARGIN = 28.0
LINE = 11.5
HEADING = 8.5
BODY = 7.6


def _text(page, x, y, value, size=BODY, bold=False):
    page.insert_text(
        (x, y), str(value), fontsize=size,
        fontname="hebo" if bold else "helv",
    )


def _right(page, x, y, value, size=BODY, bold=False):
    value = str(value)
    font = "hebo" if bold else "helv"
    width = fitz.get_text_length(value, fontname=font, fontsize=size)

    page.insert_text((x - width, y), value, fontsize=size, fontname=font)


def _fit(value, width, size=BODY):
    value = str(value)

    if fitz.get_text_length(value, fontsize=size) <= width:
        return value

    while value and fitz.get_text_length(value + "…", fontsize=size) > width:
        value = value[:-1]

    return value + "…"


# Column layout for the PDF, as (field, label, x, width, right-aligned).
PDF_COLUMNS = [
    ("invoice_no", "Invoice #", 0, 54, False),
    ("date", "Date", 56, 46, False),
    ("customer", "Customer / Address", 104, 138, False),
    ("item", "Item", 246, 110, False),
    ("qty", "Qty", 360, 26, True),
    ("bonus", "Bonus", 390, 30, True),
    ("price", "T.Price", 424, 44, True),
    ("discount", "Disc%", 472, 34, True),
    ("line_total", "Amount", 510, 54, True),
    ("invoice_total", "Inv Total", 568, 58, True),
    ("paid", "Paid", 630, 50, True),
    ("balance", "Balance", 684, 54, True),
    ("status", "Status", 742, 44, False),
]

# The rightmost column has to end inside the page, or the balance everyone
# opened the report for is the one figure that gets cut off.
assert PDF_COLUMNS[-1][2] + PDF_COLUMNS[-1][3] <= PAGE.width - 2 * MARGIN


def _new_page(document, title, subtitle):
    page = document.new_page(width=PAGE.width, height=PAGE.height)

    _text(page, MARGIN, MARGIN + 4, title, size=13, bold=True)
    _text(page, MARGIN, MARGIN + 18, subtitle, size=8)

    y = MARGIN + 36

    for _, label, x, width, right in PDF_COLUMNS:
        if right:
            _right(page, MARGIN + x + width, y, label, size=HEADING, bold=True)
        else:
            _text(page, MARGIN + x, y, label, size=HEADING, bold=True)

    page.draw_line(
        fitz.Point(MARGIN, y + 3), fitz.Point(PAGE.width - MARGIN, y + 3),
        width=0.7,
    )

    return page, y + LINE + 2


def to_pdf(invoices, title="Invoices", subtitle=""):
    """The same list as a document, for printing and handing round."""
    document = fitz.open()

    page, y = _new_page(document, title, subtitle)

    bottom = PAGE.height - MARGIN - 40

    rows = invoice_rows(invoices)

    for position, row in enumerate(rows):
        if y > bottom:
            page, y = _new_page(document, title, f"{subtitle} (continued)")

        elif row["first"] and position:
            page.draw_line(
                fitz.Point(MARGIN, y - 7),
                fitz.Point(PAGE.width - MARGIN, y - 7),
                width=0.25, color=(0.78, 0.82, 0.86),
            )

        for field, _, x, width, right in PDF_COLUMNS:
            value = row[field]

            if value == "" or value is None:
                continue

            if field in ("invoice_no", "date", "customer", "status") and not row["first"]:
                # A second line of the same invoice: the customer, the date
                # and the number are already above it, and repeating them
                # buries the items they belong to.
                continue

            if field == "date":
                value = value.strftime("%d/%m/%Y")
            elif field == "customer":
                # Name on this line, address underneath in the same column.
                _text(page, MARGIN + x, y, _fit(row["customer"], width))

                if row["address"]:
                    _text(
                        page, MARGIN + x, y + LINE - 3.5,
                        _fit(row["address"], width, size=6.4), size=6.4,
                    )
                continue
            elif field in MONEY_COLUMNS:
                value = f"{Decimal(value):,.2f}"
            elif field == "discount":
                value = f"{Decimal(value):g}%"

            if right:
                _right(page, MARGIN + x + width, y, _fit(value, width))
            else:
                _text(page, MARGIN + x, y, _fit(value, width))

        # A first line carries the address beneath it, so it needs the room.
        y += LINE + (4.0 if row["first"] and row["address"] else 0)

    figures = totals_of(invoices)

    if y > bottom:
        page, y = _new_page(document, title, f"{subtitle} (continued)")

    page.draw_line(
        fitz.Point(MARGIN, y - 4), fitz.Point(PAGE.width - MARGIN, y - 4),
        width=0.7,
    )

    y += LINE - 4

    _text(page, MARGIN, y, f"{figures['invoices']} invoice(s)", bold=True)

    for field, value in (
        ("invoice_total", figures["total"]),
        ("paid", figures["paid"]),
        ("balance", figures["balance"]),
    ):
        spec = next(c for c in PDF_COLUMNS if c[0] == field)

        _right(
            page, MARGIN + spec[2] + spec[3], y, f"{value:,.2f}", bold=True
        )

    for number, one in enumerate(document, start=1):
        _text(
            one, PAGE.width - MARGIN - 70, PAGE.height - MARGIN + 6,
            f"Page {number} of {document.page_count}", size=7,
        )

    output = document.tobytes()
    document.close()

    return output
